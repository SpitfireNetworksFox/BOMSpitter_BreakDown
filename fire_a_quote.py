# excel_to_quote_html.py
from __future__ import annotations
import argparse
import base64
import json
import re
import os
from pathlib import Path
from datetime import datetime, timedelta
from collections import OrderedDict  # ← ADDED

import numpy as np
import pandas as pd
from jinja2 import Environment, BaseLoader
from openpyxl import load_workbook
from utils.quote_notes import notes as quote_notes
from utils.client_and_account_rep import clientAndAccountReps
import utils.fire_a_quote_utils as utils

# ------------------ Config ------------------

LOGO_PATH = Path("utils/logo_small.jpg")
with open(LOGO_PATH, "rb") as f:
    logo_bytes = f.read()
LOGO_B64 = base64.b64encode(logo_bytes).decode("utf-8")

CANDIDATES = {
    "sku": ["sku", "part", "item", "product", "pn", "part number", "mpn", "partner sku"],
    "qty": ["qty", "quantity", "qnty", "q'ty"],
    "description": ["description", "desc", "product description", "item description"],
    "unit_price": ["unit price", "price", "list price", "unitprice", "price ea", "each"],
    "discount": ["discount"],
    "discount_price": ["discount price", "discounted price", "sell price", "your price", "net price"],
    "extended": ["extended", "extended price", "ext", "subtotal", "line total", "amount"],
    "notes": ["notes", "note", "comments", "line notes"],
    "category": ["product type", "category", "product category", "type"],

    "start_date": ["start date", "start", "service start", "begin", "term start"],
    "end_date": ["end date", "end", "service end", "finish", "term end"]
}


DEFAULTS = clientAndAccountReps

# ------------------ Reading strategies ------------------


def read_with_pandas_values(excel_path: Path, sheet: str) -> pd.DataFrame:
    # Pandas/openpyxl reads cached values for formulas (if present)
    return pd.read_excel(excel_path, sheet_name=sheet)


def read_with_openpyxl_cached(excel_path: Path, sheet: str) -> pd.DataFrame:
    # Explicitly read the *cached* (already-computed) values Excel stored last time it was saved.
    wb = load_workbook(excel_path, data_only=True, read_only=False)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = rows[1:]
    return pd.DataFrame(data, columns=headers)


def try_excel_recalc(excel_path: Path):
    # Optional: opens Excel (Windows/macOS) and calculates formulas
    try:
        import xlwings as xw
        app = xw.App(visible=False)  # no window
        book = xw.Book(str(excel_path))
        book.app.calculate()
        book.save()
        book.close()
        app.quit()
        return True
    except Exception:
        return False

# ------------------ Items parsing ------------------


def _fmt_date(v):
    import pandas as pd
    from datetime import datetime, date
    from numbers import Number

    # Empty/null
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    if isinstance(v, str) and not v.strip():
        return ""

    # Pandas/py datetimes
    if isinstance(v, (pd.Timestamp, datetime, date)):
        return v.strftime("%Y-%m-%d")

    # Excel serial numbers (sometimes appear as plain numbers)
    if isinstance(v, Number):
        # Try Excel-origin serial first (works for most modern workbooks)
        try:
            # Excel serial date: days since 1899-12-30 (pandas' origin='1899-12-30')
            dt = pd.to_datetime(
                v, unit="D", origin="1899-12-30", utc=False, errors="raise")
            return dt.strftime("%Y-%m-%d")
        except Exception:
            # Fallback: treat as year if reasonable, else give up to string
            pass

    # Strings → try to parse common patterns; else leave as-is
    s = str(v).strip()
    try:
        dt = pd.to_datetime(s, errors="raise", dayfirst=False)
        # If it parsed to a full timestamp, normalize to date
        if isinstance(dt, pd.Timestamp):
            return dt.strftime("%Y-%m-%d")
    except Exception:
        pass

    # Last resort: return the user-provided string unchanged
    return s


def parse_items(df_raw: pd.DataFrame) -> list[dict]:
    df = utils._norm_cols(df_raw).dropna(how="all")
    if df.empty:
        return []

    # Column picks (keeps your flexible header matching)
    col_sku = utils._pick(df, [c.lower() for c in CANDIDATES["sku"]])
    col_qty = utils._pick(df, [c.lower() for c in CANDIDATES["qty"]])
    col_desc = utils._pick(df, [c.lower() for c in CANDIDATES["description"]])

    # Treat "Your Price" (discount_price family) as the LINE TOTAL if present.
    col_your = utils._pick(df, [c.lower()
                           for c in CANDIDATES["discount_price"]])
    col_list = utils._pick(df, [c.lower() for c in CANDIDATES["unit_price"]])
    col_disc = utils._pick(df, [c.lower() for c in CANDIDATES["discount"]])

    # Start Date and End
    col_start = utils._pick(df, [c.lower() for c in CANDIDATES["start_date"]])
    col_end = utils._pick(df, [c.lower() for c in CANDIDATES["end_date"]])

    # Ignore any provided "Extended" column on purpose.
    col_note = utils._pick(df, [c.lower() for c in CANDIDATES["notes"]])
    col_cat = utils._pick(df, [c.lower() for c in CANDIDATES["category"]])

    items: list[dict] = []
    for _, r in df.iterrows():
        sku = (str(r.get(col_sku)) if col_sku else "").strip()
        desc = (str(r.get(col_desc)) if col_desc else "").strip()
        if (not sku or sku.lower() == "nan") and (not desc or desc.lower() == "nan"):
            continue

        # Skip "Totals" rows
        first_col_name = df.columns[0]
        first_val = r.get(first_col_name)
        if isinstance(first_val, str) and first_val.strip().lower() == "totals":
            continue

        # QTY (default 1 if blank/invalid)
        qty = utils._num(r.get(col_qty)) if col_qty else 1
        qty = int(qty) if qty and qty > 0 else 1

        # Primary path: "Your Price" is the LINE TOTAL
        your_price_total = utils._num(r.get(col_your)) if col_your else None

        unit_price = None
        extended = 0.0

        if your_price_total is not None and not np.isnan(your_price_total):
            # Unit Price = Your Price / QTY ; Extended = Your Price
            unit_price = (your_price_total / qty) if qty else your_price_total
            extended = your_price_total
        else:
            # Fallback path (if "Your Price" missing): use list & discount or list itself
            list_price = utils._num(r.get(col_list)) if col_list else None
            disc = utils._num(r.get(col_disc)) if col_disc else 0
            if list_price is not None:
                unit_price = list_price * (1 - (disc or 0))
                extended = unit_price * qty
            else:
                unit_price = 0.0
                extended = 0.0

        note_val = (str(r.get(col_note)).strip()
                    if col_note and pd.notna(r.get(col_note)) else "")
        category_val = (str(r.get(col_cat)).strip()
                        if col_cat and pd.notna(r.get(col_cat)) else "")

        items.append({
            "pt_sku": "" if sku.lower() == "nan" else sku,
            "qty": qty,
            "description": "" if desc.lower() == "nan" else desc,
            # Keep list_price if you want to optionally display it via --show-list-price
            "list_price": float(utils._num(r.get(col_list))) if col_list and pd.notna(r.get(col_list)) else None,
            # New: include explicit unit_price for the template
            "unit_price": float(unit_price) if unit_price is not None else 0.0,
            # Extended (subtotal) is always the line total (Your Price)
            "subtotal": float(extended),
            "notes": note_val,
            "category": category_val,
            "start_date": _fmt_date(r.get(col_start)) if col_start else "",
            "end_date":   _fmt_date(r.get(col_end)) if col_end else "",
        })
    return items

# ------------------ HTML render ------------------


def render_html(context: dict, template_html: str) -> str:
    env = Environment(loader=BaseLoader())
    return env.from_string(template_html).render(**context)

# ------------------ Main ------------------


def main():
    ap = argparse.ArgumentParser(
        description="Convert Excel Quote to HTML+PDF, or turn an existing HTML into a PDF"
    )

    # Make --excel and --usehtml mutually exclusive
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--excel", help="Path to Excel .xlsx")
    src.add_argument(
        "--usehtml", help='Path to an existing HTML file to convert directly to PDF')

    # Existing args (apply to Excel→HTML path only)
    ap.add_argument("--out", default=None,
                    help=('Output path: '
                          'Excel mode → HTML file (default: "Generated Quotes/<excel name> - <quote #>.html"); '
                          'HTML mode → PDF file (default: "<usehtml dirname>/<usehtml basename>.pdf")'))
    ap.add_argument("--items-sheet", default=None,
                    help="Items sheet name (optional; Excel mode only)")
    ap.add_argument("--header-sheet", default=None,
                    help="Header sheet name (optional; Excel mode only)")
    ap.add_argument("--round-values", choices=["True", "False"],
                    default=None, help="Round currency to whole dollars (Excel mode only)")
    ap.add_argument("--excel-recalc", action="store_true",
                    help="Open Excel via xlwings to force a recalc before reading (Excel mode only)")
    ap.add_argument("--show-list-price", choices=["True", "False"], default="False",
                    help="Include the 'List Price' column when --round-values=False (Excel mode only)")
    ap.add_argument("--usa", action="store_true",
                    help='Use the US template (utils/quote_html_template_usa.py) instead of the default (Excel mode only)')
    args = ap.parse_args()

    # ─────────────────────────────
    # Fast-path: HTML → PDF only
    # ─────────────────────────────
    if args.usehtml:
        html_path = Path(args.usehtml)
        if not html_path.exists():
            raise FileNotFoundError(html_path)

        # If --out is provided in HTML mode, treat it as the PDF output path.
        # Otherwise, default to "<usehtml>.pdf" in the same directory.
        pdf_out = Path(args.out) if args.out else html_path.with_suffix(".pdf")
        # utils.write_pdf_via_playwright() writes the PDF next to the HTML by default.
        # If it doesn't support a custom output path, we can move/rename after.
        # But first, write the PDF using the existing function:
        utils.write_pdf_via_playwright(html_path)

        # If a custom --out was provided and differs from default, rename
        default_pdf = html_path.with_suffix(".pdf")
        if pdf_out.resolve() != default_pdf.resolve():
            pdf_out.parent.mkdir(parents=True, exist_ok=True)
            default_pdf.replace(pdf_out)

        print(json.dumps({
            "mode": "html_to_pdf",
            "input_html": str(html_path.resolve()),
            "output_pdf": str(pdf_out.resolve())
        }, indent=2))
        return

    # ─────────────────────────────
    # Excel → HTML (+PDF) path
    # ─────────────────────────────
    excel_path = Path(args.excel)
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    # Choose template lazily (so imports only happen when needed)
    if args.usa:
        from utils.quote_html_template_usa import quoteTemplateHtml as TEMPLATE_HTML
    else:
        from utils.quote_html_template import quoteTemplateHtml as TEMPLATE_HTML

    # 1) (optional) force real Excel recalc
    if args.excel_recalc:
        ok = try_excel_recalc(excel_path)
        if not ok:
            print("Warning: Excel recalc via xlwings failed or not available. Continuing with cached/fallback values.")

    # Pick items sheet
    xls = pd.ExcelFile(excel_path)
    items_sheet = args.items_sheet or utils._prefer_sheet(
        xls, ["items", "quote", "lines", "sheet1", "my customer deal"]
    )

    # 2) try cached (computed) values explicitly via openpyxl
    df_cached = read_with_openpyxl_cached(excel_path, items_sheet)

    # 3) fall back to pandas
    if df_cached.empty:
        df_cached = read_with_pandas_values(excel_path, items_sheet)

    items = parse_items(df_cached)
    bom_total = float(np.nansum([i["subtotal"]
                      for i in items])) if items else 0.0

    # ---- NEW: group items by Category & compute per-category subtotals ----
    def _cat_name(v: str) -> str:
        s = (v or "").strip()
        return s if s else "Uncategorized"

    # Preserve first-seen order of categories from the incoming items list
    ordered_categories: list[str] = []
    seen = set()
    for it in items:
        c = _cat_name(it.get("category"))
        if c not in seen:
            ordered_categories.append(c)
            seen.add(c)

    groups = OrderedDict()
    for c in ordered_categories:
        groups[c] = {"category": c, "items": [], "subtotal": 0.0}

    for it in items:
        c = _cat_name(it.get("category"))
        g = groups[c]
        g["items"].append(it)
        g["subtotal"] += float(it.get("subtotal") or 0.0)

    items_by_category = list(groups.values())
    # ----------------------------------------------------------------------

    # Minimal header meta
    meta = DEFAULTS.copy()
    meta["bom_name"] = excel_path.stem.replace("_", " ")
    if args.round_values is not None:
        meta["round_values"] = args.round_values

    meta["quote_number"] = utils.pop_and_increment_quote_number()

    # Choose output path (default uses "Generated Quotes/<excel_stem> - <quote_number>.html")
    out_dir = Path("Generated Quotes")
    out_dir.mkdir(parents=True, exist_ok=True)

    default_out = out_dir / \
        f"{utils._safe_filename(excel_path.stem)} - {utils._safe_filename(meta['quote_number'])}.html"
    out_path = Path(args.out) if args.out else default_out
    out_path.parent.mkdir(parents=True, exist_ok=True)  # ensure folder exists

    # Add default notes
    if quote_notes is None:
        meta["add_notes"] = [
            "This quote is missing notes. Please contact your Spitfire representative.",
        ]
    else:
        meta["add_notes"] = quote_notes

    owner = {"name": meta["owner_name"], "email": meta["owner_email"]}
    context = {
        "img_str": LOGO_B64,
        "quote_number": meta["quote_number"],
        "date": meta["date"],
        "payment_terms": meta["payment_terms"],
        "expiry": meta["expiry"],
        "currency": meta["currency"],
        "owner_phone": meta["owner_phone"],
        "round_values": meta["round_values"],
        "sales_desk_email": meta["sales_desk_email"],
        "owner": owner,

        # Keep original list for any legacy template sections
        "items": items,

        # NEW: grouped view for category headers + subtotals in the template
        "items_by_category": items_by_category,

        # So the template can decide to show/hide List Price
        "show_list_price": args.show_list_price,

        "bom": {
            "bom_name": meta["bom_name"],
            "contact_name": meta["contact_name"],
            "comp_name": meta["comp_name"],
            "comp_address": meta["comp_address"],
            "comp_city": meta["comp_city"],
            "comp_state": meta["comp_state"],
            "comp_zip": meta["comp_zip"],
            "comp_phone": meta["comp_phone"],
            "term": meta["term"],
            "incoterms": meta["incoterms"],
            "duty": meta["duty"],
            "taxes": meta["taxes"],
            "total": bom_total,
        },
        "quote_notes": None,
        "add_notes": meta["add_notes"],
    }

    # Render HTML + write
    html = render_html(context, TEMPLATE_HTML)
    out_path.write_text(html, encoding="utf-8")

    # Create a matching PDF (same name, same directory)
    utils.write_pdf_via_playwright(out_path)

    print(json.dumps({
        "mode": "excel_to_html_pdf",
        "excel": str(excel_path.resolve()),
        "items_sheet_used": items_sheet,
        "parsed_items": len(items),
        "category_groups": len(items_by_category),  # ← helpful runtime info
        "total": round(bom_total, 2),
        "output_html": str(out_path.resolve()),
        "output_pdf": str(out_path.with_suffix('.pdf').resolve())
    }, indent=2))


if __name__ == "__main__":
    main()
